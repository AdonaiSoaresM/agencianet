from openpyxl import load_workbook
from selenium import webdriver
import sys
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
import time
import numpy as np
import os

print('Iniciando!')

def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.dirname(__file__)
    return os.path.join(base_path, relative_path)


homedir = os.getlogin()
#nota = np.array([['NF-E','NF-E'],['CT-E','CT-E'],['MDF-E','NFC-E']])
nota = ['NF-E','CT-E','NFC-E','MDF-E']
wb = load_workbook(r'C:\\Users\\' + homedir + '\\Desktop\\pyt.xlsx') #LOCAL DA PLANILHA
plan = wb.worksheets[0]
totalcnpj = 1

while ((plan.cell(totalcnpj, 1).value) != None):
    totalcnpj = totalcnpj + 1

navegador = webdriver.Chrome("C:/Users/" + os.getlogin() + "/Documents/selenium/chromedriver.exe")
navegador.implicitly_wait(30)
navegador.maximize_window()
navegador.get("https://www2.agencianet.fazenda.df.gov.br/Inicio/Restrita/")
navegador.get("https://www2.agencianet.fazenda.df.gov.br/DEC/#/")

time.sleep(5)
    
for linha in range(totalcnpj):
    cnpj = wb.worksheets[0].cell(linha+1,1).value
    dataInicial = wb.worksheets[0].cell(linha+1,2).value
    dataFim = wb.worksheets[0].cell(linha+1,3).value

    for rad in range(2): #LOOP NO TIPO EMITENTE/DESTINATARIO
        
        for tipo in range(4): #LOOP NAS NOTAS FISCAIS (SELECT)

            try: #ESPERAR A CAIXA DE CARREGAMENTO SUMIR
                element = WebDriverWait(navegador, 2000).until(
                    EC.invisibility_of_element_located((By.ID,'caixa-carregando'))
                )
                
            finally:
                    
                time.sleep(1)
                radio = navegador.find_elements_by_class_name('radio-inline')
                
                if (cnpj is None): #QUANDO ACABA OS CNPJS NA PLANILHA
                    print("CNPJ VAZIO NA LINHA " + str(linha+1) + ", FEITO!")
                    sys.exit()
                libera = tipo + rad
                
                if libera != 4: #NÃO IMPRIMIR A NOTA MDF-E PARA DESTINATARIO
                    navegador.find_element_by_id('CpfCnpj').send_keys(Keys.BACKSPACE * 3,cnpj) #CNPJ
                    navegador.find_element_by_id('DataInicio').send_keys(Keys.BACKSPACE * 3,dataInicial) #DATA INICIAL
                    dtiSite = navegador.find_element_by_id('DataInicio').get_attribute('value')
                    dtiPlan = ((dataInicial[0:2] + '/') + (dataInicial[2:4] + '/') + (dataInicial[4:8]))

                    while dtiSite != dtiPlan: #VERIFICA SE O SENDKEYS DA DATA É IGUAL AO DA PLANILHAS
                        navegador.find_element_by_id('DataInicio').click()
                        navegador.find_element_by_id('DataInicio').clear()
                        navegador.find_element_by_id('DataInicio').send_keys(Keys.BACKSPACE * 3,dataInicial)
                        #print('Data Inicial Diferente: ' + dtiSite) 
                        dtiSite = navegador.find_element_by_id('DataInicio').get_attribute('value')
                        print(dtiSite + " | " + dtiPlan)
                        #print('Data Inicial Após tentativa de correção: ' + dtiSite)
                        time.sleep(1)
                    
                    navegador.find_element_by_id('DataFim').send_keys(Keys.BACKSPACE * 3,dataFim) #DATA FIM
                    
                    dtfSite = navegador.find_element_by_id('DataFim').get_attribute('value')
                    dtfPlan = ((dataFim[0:2] + '/') + (dataFim[2:4] + '/') + (dataFim[4:8]))

                    while dtfSite != dtfPlan: #VERIFICA SE O SENDKEYS DA DATA É IGUAL AO DA PLANILHAS
                        navegador.find_element_by_id('DataFim').click()
                        navegador.find_element_by_id('DataFim').clear()
                        navegador.find_element_by_id('DataFim').send_keys(Keys.BACKSPACE * 3,dataFim)
                        #print('Data Final Diferente: ' + dtfSite) 
                        dtfSite = navegador.find_element_by_id('DataFim').get_attribute('value')
                        #print('Data Final Após tentativa de correção: ' + dtfSite)
                        time.sleep(1)
                    
                    select = Select(navegador.find_element_by_name('Tipo')) #SELECT
                    
                    #print("rad: " + str(rad) + " | tipo: " + str(tipo))
                    if(rad == 0):
                        print("Emitente")
                    else:
                        print("Destinatário")
                    print("CNPJ: " + cnpj + " | " + nota[tipo])
                    select.select_by_visible_text(nota[tipo]) #VALOR NO SELECT
                    radio[rad].click() #CLICA NO RADIO
                    WebDriverWait(navegador, 30)
                    time.sleep(2)
                    butao = navegador.find_elements_by_class_name('btn')
                    butao[2].click()
                    print("Clicou em solicitar")
                    
                    time.sleep(3)

                    try: #ESPERANDO O ARQUIVO SOLICITADO APARECER NA TABELA
                        element = WebDriverWait(navegador, 2000).until(
                            EC.invisibility_of_element_located((By.ID,'caixa-carregando'))
                        )
                        
                    finally:

                        time.sleep(3)
                        navegador.find_element_by_css_selector("img[src*='atualiza']").click()
                        print("Clicou em atualizar")

                        time.sleep(3)
                        try: #ESPERAR A TELA DE CARREGAMENTO
                            element = WebDriverWait(navegador, 2000).until(
                                EC.invisibility_of_element_located((By.ID,'caixa-carregando'))
                            )
                            
                        finally: #CLICAR PARA BAIXAR O ARQUIVO

                            time.sleep(2)
                            navegador.find_element_by_css_selector("img[src*='baixar']").click()
                            print("Clicou em baixar")
                            time.sleep(3)
                            navegador.find_element_by_css_selector("img[src*='cancela']").click()
                            print("Clicou em excluir")
                            time.sleep(5)
                            try: #ESPERAR A TELA DE CARREGAMENTO
                                element = WebDriverWait(navegador, 2000).until(
                                    EC.invisibility_of_element_located((By.ID,'caixa-carregando'))
                                )
                                
                            finally: #LIMPA
                                time.sleep(3)
                                butao[3].click()
                                print("Clicou em Limpar")
                                print("____________________________________________")

                
    

